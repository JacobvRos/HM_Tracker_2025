"""
Rat Hex Maze Preprocessing — Step 1 + Step 2 with Homeboxes

Finds every .xlsx in EXCEL_FOLDER, processes all rows, and saves one
results Excel file per input file into OUTPUT_FOLDER.

The output is a copy of the original file with the computed columns
appended to the right of the data sheet. All other worksheets and all
existing cell colours are preserved. Only the newly added cells in
flagged rows are coloured red.

Computed columns added:
  Step 1 — shortest_path, eat_on_1_encounter, n_nodes_visited, food_reached,
            dist_tra, dt_rel_sp, dt_min_sp, dir_run_mat_perf,
            node_choices_binary, perc_corr_choices
  Step 2 — node_island_in, island_short_path, island_dt_traveled, perf_in_island
  Errors — flag  (non-empty = row skipped, new cells highlighted red)

Requirements: pip install networkx pandas openpyxl
"""

import argparse
import glob
import os
import shutil
import numpy as np
import pandas as pd
import networkx as nx
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── Parse folder from command line ───────────────────────────────────────────
parser = argparse.ArgumentParser(description='Rat Hex Maze Analysis')
parser.add_argument('--input_folder',  '-i', required=True,
                    help='Input folder (ip[n]) containing .xlsx files')
parser.add_argument('--output_folder', '-o', required=True,
                    help='Output folder (op[n]) where results are saved')
args = parser.parse_args()

FOLDER_PAIRS = [(args.input_folder, args.output_folder)]
print(f'Input:  {args.input_folder}')
print(f'Output: {args.output_folder}')

# ── Graph: 96 maze nodes + 2 homeboxes (501, 502) ────────────────────────────
NDS = (list(range(101, 125)) + list(range(201, 225)) +
       list(range(301, 325)) + list(range(401, 425)) + [501, 502])

_s1 = [1,1,2,3,3,4,5,6,6,7,8,8,9,10,10,11,12,13,14,14,15,16,16,17,18,18,20,21,22,23]
_t1 = [2,7,3,4,9,5,11,13,7,8,9,15,10,11,17,12,19,14,20,15,16,22,17,18,19,24,21,22,23,24]
_between = [(24,25),(44,53),(72,73),(21,50),(47,76),(21,97),(47,98),(50,97),(76,98)]

def _build_graph():
    G = nx.Graph()
    G.add_nodes_from(NDS)
    for offset in [0, 24, 48, 72]:
        for s, t in zip(_s1, _t1):
            G.add_edge(NDS[offset + s - 1], NDS[offset + t - 1])
    for s, t in _between:
        G.add_edge(NDS[s - 1], NDS[t - 1])
    return G

G     = _build_graph()
_DIST = dict(nx.all_pairs_shortest_path_length(G))

RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

OUTPUT_COLS = [
    'shortest_path', 'eat_on_1_encounter', 'n_nodes_visited', 'food_reached',
    'dist_tra', 'dt_rel_sp', 'dt_min_sp', 'dir_run_mat_perf',
    'node_choices_binary', 'perc_correct_choices',
    'isl_node_in', 'isl_short_path', 'isl_dt_trav', 'perf_in_island',
    'flag',
]

def _pick_sheet(xl):
    for name in xl.sheet_names:
        if name.lower() == 'raw':
            return name
    return xl.sheet_names[0]

def _compute(exc_path):
    """Read the data sheet and compute all new columns. Returns (df, sheet_name)."""
    xl    = pd.ExcelFile(exc_path)
    sheet = _pick_sheet(xl)
    df    = pd.read_excel(exc_path, sheet_name=sheet, header=0)
    print(f'  Sheet: "{sheet}"  |  {len(df)} rows')

    for col in OUTPUT_COLS:
        df[col] = np.nan if col not in ('flag', 'node_choices_binary') else ''

    for i, row in df.iterrows():
        if pd.isna(row['path_to_reach']) or str(row['path_to_reach']).strip() == '':
            continue

        try:
            path       = [int(float(x)) for x in str(row['path_to_reach']).split(',') if x.strip()]
            start_node = int(row['start_node_n'])
            goal_node  = int(row['goal_node_n'])

            if path[0] != start_node:
                raise ValueError(f'start_node_n ({start_node}) != first path node ({path[0]})')

            unknown = [n for n in path if n not in _DIST]
            if unknown:
                raise ValueError(f'unknown node(s) in path: {unknown}')

            n_nodes      = len(path)
            food_reached = goal_node in path[-2:]
            shortest     = _DIST[start_node][goal_node]
            dist_tra     = (n_nodes - 1) if food_reached else 99

            # ── Step 1 ───────────────────────────────────────────────────────
            df.at[i, 'shortest_path']      = shortest
            df.at[i, 'eat_on_1_encounter'] = int(path[-1] == goal_node)
            df.at[i, 'n_nodes_visited']    = n_nodes
            df.at[i, 'food_reached']       = int(food_reached)
            df.at[i, 'dist_tra']           = dist_tra
            df.at[i, 'dt_rel_sp']          = dist_tra / shortest if shortest > 0 else np.nan
            df.at[i, 'dt_min_sp']          = dist_tra - shortest
            df.at[i, 'dir_run_mat_perf']   = int(food_reached and dist_tra == shortest)

            choices = []
            for iNode in range(n_nodes - 1):
                curr       = path[iNode]
                next_n     = path[iNode + 1]
                neighbours = list(G.neighbors(curr))
                if next_n not in neighbours:
                    choices.append(0)
                    continue
                min_sp = min(_DIST[nb][goal_node] for nb in neighbours)
                choices.append(int(_DIST[next_n][goal_node] == min_sp))

            df.at[i, 'node_choices_binary']  = ','.join(str(c) for c in choices)
            df.at[i, 'perc_correct_choices'] = (
                (sum(choices) * 100) / (n_nodes - 1) if n_nodes > 1 else np.nan
            )

            # ── Step 2 (skipped if trial is excluded) ────────────────────────
            if row['exclude_trial'] != 0:
                continue

            diffs         = [abs(path[j + 1] - path[j]) for j in range(n_nodes - 1)]
            enter_indices = [j for j, d in enumerate(diffs) if d >= 50]

            if enter_indices:
                index_enter    = enter_indices[-1]
                node_island_in = path[index_enter]
                island_sp      = _DIST[node_island_in][goal_node] + 1
                island_dt      = len(path[index_enter:])
                df.at[i, 'isl_node_in']    = node_island_in
                df.at[i, 'isl_short_path'] = island_sp
                df.at[i, 'isl_dt_trav']    = island_dt
                df.at[i, 'perf_in_island'] = island_dt / island_sp

        except Exception as e:
            df.at[i, 'flag'] = str(e)
            print(f'  Flagged row {i + 2}: {e}')

    return df, sheet

def _save(df, sheet, exc_path, out_path):
    """Copy the original file, find existing column headers by name, and write
    computed values into them. Overwrites any existing values. All other sheets,
    formatting, and cell colours are untouched. Only the newly written cells in
    flagged rows are coloured red."""
    shutil.copy2(exc_path, out_path)

    wb = load_workbook(out_path)
    ws = wb[sheet]

    # Build a map of header name → column index (1-based) from row 1
    header_to_col = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }

    # Columns that may appear under different names in different Excel files
    ALIASES = {
        'perc_correct_choices': ['perc_correct_choices', 'perc_corr_choices'],
        'isl_node_in':          ['isl_node_in',          'node_island_in'],
        'isl_short_path':       ['isl_short_path',       'island_short_path'],
        'isl_dt_trav':          ['isl_dt_trav',          'island_dt_traveled'],
    }

    # For any OUTPUT_COL not already in the sheet, add it at the end
    next_new_col = ws.max_column + 1
    col_map = {}
    for col_name in OUTPUT_COLS:
        candidates = ALIASES.get(col_name, [col_name])
        matched = next((c for c in candidates if c in header_to_col), None)
        if matched:
            col_map[col_name] = header_to_col[matched]
        else:
            print(f'  WARNING: "{col_name}" not found in sheet — adding as new column')
            ws.cell(row=1, column=next_new_col, value=col_name)
            col_map[col_name] = next_new_col
            next_new_col += 1

    # Clear all existing values in the output columns (rows 2 onwards)
    for col_idx in col_map.values():
        for excel_row in range(2, ws.max_row + 1):
            ws.cell(row=excel_row, column=col_idx).value = None

    # Write values into the correct columns
    for i, row in df.iterrows():
        excel_row  = i + 2      # pandas index 0 → Excel row 2 (row 1 is header)
        is_flagged = bool(row['flag'])

        for col_name, col_idx in col_map.items():
            val = row[col_name]
            if isinstance(val, float) and np.isnan(val):
                val = None
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            if is_flagged:
                cell.fill = RED_FILL

    wb.save(out_path)

# ── Run over every ip/op pair ─────────────────────────────────────────────────
for ip_folder, op_folder in FOLDER_PAIRS:
    os.makedirs(op_folder, exist_ok=True)
    excel_files = glob.glob(os.path.join(ip_folder, '*.xlsx'))

    if not excel_files:
        print(f'No .xlsx files found in {ip_folder}')
        continue

    print(f'\n── {ip_folder} → {op_folder}')
    for exc_path in excel_files:
        fname = os.path.basename(exc_path)
        print(f'\n  Processing: {fname}')
        df, sheet = _compute(exc_path)
        out_path  = os.path.join(op_folder, fname.replace('.xlsx', '_results.xlsx'))
        _save(df, sheet, exc_path, out_path)
        n_flagged = (df['flag'] != '').sum()
        print(f'  Saved → {out_path}  ({n_flagged} row(s) flagged red)')

print('\nAll files done.')
